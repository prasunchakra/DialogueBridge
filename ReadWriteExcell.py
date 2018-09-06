import openpyxl

workbook=openpyxl.load_workbook('Sample.xlsx')
print (workbook.sheetnames)
print ("Active Sheet: " + workbook.active.title) # Sheet that opens first

sheet = workbook['Department']

print (sheet['A3'].value)
print (sheet.cell(2,1).value)
print (sheet.max_row)
print (sheet.max_column)


for r in sheet['A1':'B5']:
    for c in r:
        print (c.value, end=' ')
    print('\n')

workbook.create_sheet(title="Department2")
sheet2 = workbook["Department2"]
workbook.remove(workbook["Employee"])
workbook.save("Sample2.xlsx")